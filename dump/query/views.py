from django.shortcuts import render
from .models import *
from django.http import HttpResponse


# Create your views here.
def csv(request):
    import csv
    from django.utils.encoding import smart_str
    queryset=MyModel.objects.all()#select queryset
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Title"),
        smart_str(u"Description"),
    ])
    for obj in queryset:
        writer.writerow([
        smart_str(obj.pk),
        smart_str(obj.title),
        smart_str(obj.description),
        ])
    return response


def xls(request):
    import xlwt
    queryset=MyModel.objects.all()#select queryset
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("MyModel")
    
    row_num = 0
    
    columns = [
        (u"ID", 2000),
        (u"Title", 6000),
        (u"Description", 8000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.title,
            obj.description,
        ]
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response

def xlsx(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    queryset=MyModel.objects.all()#select queryset
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.title,
            obj.description,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            #c.style.alignment.wrap_text = True

    wb.save(response)
    return response


